# -*- coding: utf-8 -*-
"""
투석 식단 5대 영양소 조정 레버 (규칙 기반 후처리) — 한 끼(끼니) 단위
단위: 한 끼 = 5메뉴(밥-국-찬-찬-찬). 기준 = 1일 기준 ÷ 3. 유저 지정메뉴(anchor)는 후순위로 보존.
expand() 단계 재료 정규화: 잡곡→백미, 육수(멸치·다시마·사골·시판)→맹물(육수 전면 금지).
- 나트륨: 첨가염(조미료+젓갈)만 소금1g(393mg)까지 적응적 축소. 자연염·김치절임염은 판정 제외.
- 칼륨:  상한 초과 시 기여도(양xK) 큰 재료 -> 같은군 저칼륨 교체(인 가드레일). 비앵커 우선.
- 인:    상한 초과 시 (1)비앵커 저인 교체[가공→생것, 단백질75%유지] (2)비앵커 양↓ (3)앵커 양↓ (4)실패경고.
- 단백질: 범위 벗어나면 단백질찬(메뉴 통째) 비율 스케일. 비앵커 우선.
- 열량:  초과분만큼 밥계열 덜어냄(하한0.4), 부족분은 밥 증량(최대1.3)+기름 보충. 항상 마지막.
순서: 김치→나트륨→가공품 → [칼륨→인→단백→나트륨재적용→열량]×2.

[2026-07-27 최종 채택 확정] 두부·콩류 앵커 전용 조건부 경로 반영.
- 목적: 두부·콩류 주찬은 phosphorus 판정(Peff)과 protein 레버의 증량이 서로를 되풀이해서
  깨뜨리는 충돌(phosphorus↔protein 핑퐁)이 다른 앵커보다 뚜렷해, 그 앵커에서만 인 판정을
  raw P로 통일하고 단백질·열량 증량에 raw P 예산 상한(90%)을 걸어 이 충돌을 완화한다.
- 적용범위: adjust()가 주찬(menus[2]) 또는 지정 앵커의 정체성 재료 group=='두류'일 때만
  (`_plant_protein_path_needed()`, 메뉴명 하드코딩 없음) lever_phosphorus_rawP/
  lever_protein_capped/lever_calorie_capped 경로를 타고, 그 외(생선구이·육류 등 비두류
  앵커·랜덤모드 포함)는 기존 lever_phosphorus/lever_protein/lever_calorie를 완전히 그대로
  쓴다 — 아키텍처·게이트 순서·기준값 변경 없음, 규칙기반 레버 내부의 조건부 분기 하나만 추가.
- 검증: service_rollout_verification_FOOK.py로 OLD(수정 전 백업)와 3,600건 paired 비교
  — 생선구이·육류 각 1,200건 전부 완전동일(무회귀), 두부·콩류만 생성성공률 94→100%·
  후보0개율 6→0%·비현실재료량 46→20% 개선. 상세: service_rollout_verification_report.md.
- lever_potassium/lever_sodium/lever_sodium_extra/lever_kimchi/add_oil/add_snack과
  pass1·pass2 나트륨 재검증은 전부 무변경(두 경로 공통). S1(pre-loop 나트륨 제거)은
  이번 확정에 포함하지 않음(별도 미채택 사안).
- 롤백: FOOK_adjust_levers.py.bak_before_rawP_tofu_path_20260727 로 파일 교체.

[2026-08-11 갱신] 위는 두부·콩류 전용 조건부 경로가 왜/어떻게 생겼는지의 트러블슈팅
기록으로 그대로 남겨둔다(그 조건부 경로 자체, 즉 _plant_protein_path_needed 분기와
lever_protein_capped/lever_calorie_capped의 단백질·열량 상한은 지금도 유지). 다만
"그 외(두부·콩류 아닌) 앵커는 기존 lever_phosphorus를 그대로 쓴다"는 설명은 더 이상
현재 상태가 아니다 — 전체 서비스 감사에서 일반 lever_phosphorus()도 수렴판정·후보선택을
Peff 기준으로 하고 있어 같은 phosphorus↔protein 불일치를 안고 있던 게 드러났고, 이제는
모든 경로(두부·콩류 포함)의 인 hard constraint를 raw P로 통일했다. lever_phosphorus_rawP는
더 이상 별도 로직이 아니라 lever_phosphorus에 위임하는 얇은 래퍼이고, Peff는 어디서도
하드 기준으로 쓰이지 않고 대체재 후보 선택 시의 소프트 타이브레이크 신호로만 남았다.
"""
import openpyxl, glob, os, csv, json
from collections import Counter, defaultdict

REPO = 'Exploiting-Food-Embeddings-for-Ingredient-Substitution-master/Exploiting-Food-Embeddings-for-Ingredient-Substitution-master'
# 열량 레버가 양을 조절하는 "주식(밥계열)" 메뉴.
# 하드코딩 8개만 두면 Class가 밥/일품밥/면인 31개 중 28개(카레라이스·짜장밥·우동·칼국수 등)가 빠져
# 그 끼는 줄일 대상이 없어 열량 초과를 못 고쳤음 -> load_all()에서 Class 기준으로 자동 확장.
RICE = {'백미밥', '잡곡밥', '현미밥', '흑미밥', '보리밥', '오곡밥', '콩밥', '찰밥'}
RICE_CLASSES = {'밥', '일품밥', '면'}
MENU_CLASS = {}    # 메뉴 -> Class (load_all에서 채움)
PROT_GROUPS = {'육류 및 그 제품', '어패류 및 그 제품', '난류', '두류'}
NUT = None
SUBS_P = {}   # 인 전용 대체 사전 (load_all에서 채움). 칼륨 subs와 분리.
SWAP_LOG = []   # (menu, old_ing, new_ing, nutrient, old_val, new_val) — adjust()마다 초기화

# 나트륨 기준: 자연재료 나트륨은 불가피 → 조미료(첨가염)만 따져 소금 1g(=Na 393mg)/끼.
SALT_MG = 393        # 소금 1g = Na 393mg. 끼당 조미료(첨가염) 나트륨 상한 = 판정 기준.
NA_TOTAL_CEIL = 786  # 총 나트륨 위생 상한(소금2g): 대조군 채점 등 참고용.
# 끼니 총 나트륨 목표(첨가염+음식자체 전부). 임상 지침 하루 소금 5g ÷ 3끼 ≈ 655mg.
# 넘으면 lever_sodium_extra가 고나트륨 메뉴를 통째로 줄여 이 안에 맞춘다(끼니 단위라 아침·점심도 적용).
NA_TOTAL_MEAL = 655
NA_TOTAL_WARN = 786  # 줄여도 이 값(소금2g)을 넘으면 그 끼니에 '나트륨 높음' 경고 (못 줄이는 미역국 등)
SEASON_FLOOR = 0.3   # 조미료 축소 하한(이하로는 안 줄임 = 맛 보존)


# 첨가염(판정·축소 대상) = 조미료류 + 젓갈. 젓갈(액젓·새우젓)은 group이 '어패류'라 조미료류에
# 안 잡히지만 Na 8000+/100g의 순수 조미 목적 첨가염이므로 조미료처럼 취급.
# (김치 절임염은 '채소류'라 여기 안 듦 → 의도적 제외. lever_sodium이 김치찌개 속 김치까지 줄이는 부작용.)
def is_seasoning(i):
    # 첨가염 = 조미료류 + 젓갈(액젓·새우젓). 젓갈은 group이 '어패류'라 조미료류에 안 잡히지만
    # Na 8000+/100g의 순수 조미 목적 첨가염이므로 포함.
    # 김치 절임염(채소류)은 일부러 제외: 넣으면 lever_sodium이 김치찌개·김치볶음밥 속 김치를 줄여
    # 그 음식들의 인·열량 균형을 해치는데, 나트륨은 이미 99% 통과라 이득 없음(부작용만). A/B로 확인.
    # 주의: 이 함수는 "조미료류 식품군 전체" 여부를 가리키는 넓은 판정이라, K레버가
    # "이건 나트륨 레버 소관이니 안 건드림" 표시(skip_seasoning)로도 같이 쓴다. 나트륨 축소
    # 대상만 좁히려면 is_salty_seasoning()을 쓸 것 — is_seasoning() 자체의 의미는 바꾸지 않는다.
    return i['group'] == '조미료류' or '젓' in i['ing']


# 나트륨 기여가 실질적인 조미료만 축소 대상으로 삼는다 (2026-07-24, 사용자 요청 — 식초·후추처럼
# 나트륨이 거의 없는 조미료까지 lever_sodium이 비율로 같이 줄이고 있던 문제).
# 실측(FOOK_ingredients_kor.csv, 조미료류 49종): 소금류·간장류·된장·고추장·굴소스·미소·각종소스·
# 케첩·페이스트류는 Na 240~40,626이고, 그 아래(맛술 32부터 식초·후추·고춧가루·향신료가루 전부 32
# 이하)로 뚜렷한 간극이 있어 100을 경계로 삼음. 식초·후추·설탕·참기름·참깨 등은 Na가 거의 0이라
# 줄여도 나트륨 목표치엔 기여가 없고, 괜히 양만 줄어 맛만 흐트러짐.
NA_SEASONING_MIN = 100  # mg/100g. 이 미만인 조미료류(식초·후추·설탕 등)는 나트륨 축소 대상에서 제외.


def is_salty_seasoning(i):
    """나트륨 축소(lever_sodium)의 실제 대상 — is_seasoning()이면서 나트륨이 실질적인 것만.
    K레버의 skip_seasoning처럼 '조미료류 전체를 건드리지 않는' 다른 용도에는 is_seasoning()을 그대로 쓴다."""
    return is_seasoning(i) and (i['Na'] or 0) >= NA_SEASONING_MIN


def season_na(inst):
    # 첨가염(나트륨이 실질적인 조미료+젓갈)에서 온 나트륨만 합산 (자연재료·기타 가공품·식초/후추류 제외).
    return sum(i['amt'] / 100 * i['Na'] for i in inst if is_salty_seasoning(i))

# 나트륨 레버: 김치 반찬 → 저염김치 풀에서 로테이션 치환
# 피클류는 소금 없이 식초·설탕만 써서 첨가염이 거의 0 → 저나트륨·저칼륨·저인 (투석 반찬에 이상적).
LOWNA_POOL = ['저염오이김치', '저염물김치',
              '비트무피클', '파프리카피클', '새송이피클', '오이피클']
LOWNA_RECIPES = {
    '저염오이김치': [('오이, 다다기, 생것', 50), ('생강, 뿌리줄기, 생것', 1), ('물', 30),
                     ('참기름', 1.3), ('식초, 양조', 14), ('설탕, 백설탕', 3),
                     ('소금, 정제염', 0.5), ('고추기름', 6)],
    '저염물김치': [('배추, 생것', 20), ('배, 신고, 생것', 5), ('순무, 뿌리, 생것', 5),
                   ('양파, 생것', 5), ('고춧가루', 3), ('파, 쪽파, 생것', 3),
                   ('마늘, 생것', 0.2), ('생강, 뿌리줄기, 생것', 0.2)],
    # 아래 4종: 재료를 식약청 DB명으로 매핑 (파프리카 초록은 DB에 없어 빨강+노랑으로, 새송이=큰느타리버섯)
    '비트무피클': [('무, 조선무, 생것', 30), ('비트, 뿌리, 생것', 5), ('물', 15),
                   ('식초, 양조', 3), ('설탕, 백설탕', 1.5), ('후추, 검은색, 가루', 0.2)],
    '파프리카피클': [('파프리카, 빨간색, 생것', 20), ('파프리카, 노란색, 생것', 20), ('물', 15),
                     ('식초, 양조', 3), ('설탕, 백설탕', 1.5), ('후추, 검은색, 가루', 0.2)],
    '새송이피클': [('큰느타리버섯(새송이버섯), 생것', 30), ('물', 15),
                   ('식초, 양조', 3), ('설탕, 백설탕', 1.5), ('후추, 검은색, 가루', 0.2)],
    '오이피클': [('오이, 다다기, 생것', 30), ('물', 15),
                 ('식초, 양조', 3), ('설탕, 백설탕', 1.5), ('후추, 검은색, 가루', 0.2)],
}
DISH_WORDS = ('볶음', '찌개', '국', '전', '밥', '조림', '찜', '덮밥', '비빔', '만두', '탕', '죽', '튀김', '무침', '나물')
KIMCHI_SIDES = set()
ROT = [0]


def num(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s.startswith('(') and s.endswith(')'):
        s = s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


def load_all():
    wb = openpyxl.load_workbook('식약청_영양성분10.4(수정).xlsx', read_only=True)
    ws = wb.active
    ing_nut = {}
    base_fresh = defaultdict(list)
    for row in ws.iter_rows(min_row=4, values_only=True):
        nm = row[3]
        if not (nm and str(nm).strip()):
            continue
        nm = str(nm).strip()
        d = {'E': num(row[5]), 'protein': num(row[7]), 'P': num(row[24]),
             'K': num(row[25]), 'Na': num(row[26]),
             'group': (str(row[2]).strip() if row[2] else '')}
        ing_nut[nm] = d
        if '생것' in nm and d['P'] is not None:
            base_fresh[nm.split(',')[0].strip()].append((d['P'], nm))
    wb.close()
    for b in base_fresh:
        base_fresh[b].sort()

    fw = openpyxl.load_workbook('data/FOOK_음식명_재료_조리방법.xlsx', read_only=True)
    fs = fw['음식명_재료_조리방법']
    ing2kw = {}
    kw2ing = defaultdict(Counter)
    for row in fs.iter_rows(min_row=2, values_only=True):
        ing, kw = row[1], row[4]
        if ing and str(ing).strip() and kw and str(kw).strip():
            ing2kw[str(ing).strip()] = str(kw).strip()
            kw2ing[str(kw).strip()][str(ing).strip()] += 1
    fw.close()
    kw_rep = {k: c.most_common(1)[0][0] for k, c in kw2ing.items()}

    sub_raw = json.load(open(REPO + '/substitute_by_ingredient.json', encoding='utf-8'))
    subs = {}
    for o, lst in sub_raw.items():
        names = [e if isinstance(e, str) else (e.get('substitute') or e.get('name')) for e in lst]
        subs[o] = [n for n in names if n]

    # 인 전용 대체 사전(P/단백질 비율 기반, apply_phosphorus_filter.py 산출). 칼륨 사전과 분리.
    global SUBS_P
    SUBS_P = {}
    try:
        p_raw = json.load(open(REPO + '/substitute_by_ingredient_P.json', encoding='utf-8'))
        for o, lst in p_raw.items():
            SUBS_P[o] = [e.get('substitute') for e in lst if e.get('substitute')]
    except OSError:
        pass   # P 사전 없으면 인 대체는 비활성(양 감소로만 대응)

    files = sorted([f for f in glob.glob('data/*월_raw.xlsx') if not os.path.basename(f).startswith('~$')],
                   key=lambda p: int(os.path.basename(p).split('월')[0]))
    comp = defaultdict(Counter)
    compdata = {}
    def namt(a):
        if a is None:
            return ''
        if isinstance(a, float) and a.is_integer():
            a = int(a)
        return str(a).strip()
    for f in files:
        w = openpyxl.load_workbook(f, read_only=True)
        s = w.active
        cur = None
        for row in s.iter_rows(min_row=5, values_only=True):
            dish, ing, amt = row[2], row[3], row[4]
            if dish and str(dish).strip():
                if cur and cur['ings']:
                    key = tuple(cur['ings'])
                    comp[cur['name']][key] += 1
                    compdata[(cur['name'], key)] = cur['ings']
                cur = {'name': str(dish).strip(), 'ings': []}
            if ing and str(ing).strip() and cur is not None:
                cur['ings'].append((str(ing).strip(), namt(amt)))
        if cur and cur['ings']:
            key = tuple(cur['ings'])
            comp[cur['name']][key] += 1
            compdata[(cur['name'], key)] = cur['ings']
        w.close()
    master = {n: compdata[(n, c.most_common(1)[0][0])] for n, c in comp.items()}
    # 메뉴 Class 로드 -> 밥계열(RICE) 자동 확장. 군 메뉴(260)는 여기 없으므로 app_core가 추가로 등록.
    try:
        import csv as _csv
        with open('data/FOOK_nutrition.csv', encoding='utf-8-sig') as _f:
            for _r in _csv.DictReader(_f):
                MENU_CLASS[_r['name']] = _r['Class']
        RICE.update(n for n, c in MENU_CLASS.items() if c in RICE_CLASSES)
    except (OSError, KeyError):
        pass   # Class 정보 없으면 하드코딩 8개로 폴백
    # 저염김치 풀을 마스터에 추가
    for name, ings in LOWNA_RECIPES.items():
        master[name] = [(ing, str(amt)) for ing, amt in ings]
    # 추가 간식(가래떡·절편·식빵·찹쌀떡)을 마스터+Class에 등록 → _build_snack_pool이 편입
    for name, ings in EXTRA_SNACK_RECIPES.items():
        master[name] = [(ing, str(amt)) for ing, amt in ings]
        MENU_CLASS.setdefault(name, '곡류(간식)')
    # 저칼륨 과일 간식(임상 권장 1회 섭취량, 2026-07-23). 기존 raw데이터의 단감(100g)·딸기(150g)는
    # 일반 인구 기준 분량이라 칼륨 임계치를 넘어 여기서 안전 분량으로 덮어씀(raw 원본은 안 건드림).
    for name, ings in EXTRA_FRUIT_SNACK_RECIPES.items():
        master[name] = [(ing, str(amt)) for ing, amt in ings]
        MENU_CLASS[name] = '과일(간식)'
    # 원본 raw데이터 오류 수정(2026-07-24, 사용자 확인). 다수결 병합이 틀린 값을 채택한 경우 —
    # 너비아니구이가 5·6·7·8·9·10·12월에 "미트볼, 냉동"으로 반복 입력돼(조리법 문구는 "너비아니를
    # 올려 굽는다"인데 재료만 미트볼 — 복붙 실수) 다수결로 최종 채택돼 있었음. 3·9·12월에만 진짜
    # 너비아니(소고기) 레시피가 존재, 그중 9월판(정통 방식: 얇게 썰어 두드리고 배즙양념)을 채택.
    for name, ings in RECIPE_FIX_OVERRIDE.items():
        master[name] = [(ing, str(amt)) for ing, amt in ings]
    # 참깨·들깨·견과류·건포도는 요리 재료에서 제거(2026-07-24, 사용자 확정).
    # 참깨(138개 메뉴, 대부분 1g 미만 고명)·견과류/건포도(11개, 8.9g 내외 고명)는 전량 제거.
    # 들깨(33개 메뉴, 요리명에 '들깨'가 든 것 포함)는 실측 결과 전부 다른 재료(버섯·고사리·무·
    # 브로콜리·죽순 등)가 충분해서(콩가루국처럼 그 재료 자체가 요리 전부인 경우 없음) 메뉴는
    # 그대로 두고 들깨만 뺀다. 완제품 이름에 우연히 겹치는 글자(예: '과자,크래커,땅콩샌드')까지
    # 걸릴 수 있으나, 그런 메뉴는 이미 SNACK_EXCLUDE_KW로 간식에서 빠져있어 실사용에 영향 없음.
    GARNISH_STRIP_KW = ('참깨', '들깨', '호두', '아몬드', '잣,', '땅콩', '건포도')
    for name in list(master.keys()):
        master[name] = [(ing, amt) for ing, amt in master[name]
                         if not any(kw in ing for kw in GARNISH_STRIP_KW)]
    # 김치 반찬(치환 대상) 자동 식별
    KIMCHI_SIDES.clear()
    for name in master:
        if name in LOWNA_POOL or any(w in name for w in DISH_WORDS):
            continue
        if (name.endswith('김치') or '깍두기' in name or '소박이' in name
                or name.endswith('겉절이') or name.endswith('물김치')):
            KIMCHI_SIDES.add(name)
    return ing_nut, base_fresh, ing2kw, kw_rep, subs, master


# 잡곡 → 백미 재료 대체. 투석 환자는 잡곡의 인/칼륨을 피해야 함(임상 표준: "잡곡밥보다 쌀밥").
# 방식 = 재료 치환(메뉴 통짜 교체 아님): 밥계열 메뉴 안의 잡곡 곡물(수수·팥·기장·흑미·현미·보리…)을
# 백미로 바꾸고 양은 유지. 예) 오곡밥의 팥(P426)·수수(P302)·기장(P226)→백미(P98) => 인 148→약88.
# 멥쌀·흰찹쌀은 이미 백미계열이라 제외. 밥계열 메뉴에서만 적용(팥죽 등 오작동 방지).
WHITE_RICE = '백미밥'
WHITE_RICE_ING = '멥쌀, 백미, 생것'
JAPGOK_RICE = {'오곡밥', '잡곡밥', '흑미밥', '현미밥', '보리밥', '콩밥', '찰밥', '수수밥', '기장밥', '차조밥'}
GRAIN_KW = ('수수', '기장', '흑미', '현미', '보리', '귀리', '잡곡', '팥', '찹쌀', '차조', '조,')  # 백미로 바꿀 잡곡 곡물 (멥쌀만 이미 백미라 제외)


def is_japgok_grain(ing):
    return any(k in ing for k in GRAIN_KW)


# ── 육수 전면 금지 (임상 지침) ──────────────────────────────
# 투석 식단은 육수를 아예 내지 않고 '맹물'로 끓인다. 다시마·표고·채수·사골·멸치 육수 모두 금지
# (나트륨·칼륨·인이 국물로 빠져나와 그대로 섭취됨). '끓였다 물 버리고 다시 끓이기'도 금지.
# → 육수로 쓰이는 재료는 계산·조리 모두 맹물(영양 0)로 대체한다. recipe_editor도 같은 안내를 함.
BROTH_NAME_KW = ('육수', '사골', '곰탕', '장국', '채수', '뼈국물', '우족', '족국물',
                 '인스턴트', '코인', '다시다', '과립', '스톡', '요리수', '폴리인산')  # 이름이 곧 육수
BROTH_MAX_G = 8.0          # 멸치·다시마가 이 미만이면 '육수용'으로 소량 넣은 것 → 맹물. 이상이면 섭취용(멸치볶음)
BROTH_ING = ('멸치', '다시마')      # 소량이면 육수용으로 판정


def is_broth(ing, amt):
    """육수(국물)로 쓰는 재료인가 → 맹물로 대체(육수 전면 금지).
    (1) 이름 자체가 육수/사골/곰탕/장국/채수/시판육수 (2) 멸치·다시마를 소량(육수용)으로 넣은 것."""
    if any(k in ing for k in BROTH_NAME_KW):
        return True
    return any(k in ing for k in BROTH_ING) and amt < BROTH_MAX_G


def expand(menus):
    ing_nut = NUT[0]; master = NUT[5]
    inst = []
    for mn in menus:
        rice_dish = mn in RICE                                   # 밥계열 메뉴에서만 곡물 치환
        disp = WHITE_RICE if mn in JAPGOK_RICE else mn           # 순수 잡곡밥은 재료가 다 백미 → 백미밥 표시
        for ing, amt in master.get(mn, []):
            a = num(amt)
            if a is None:
                continue
            if is_broth(ing, a):                                 # 육수용 재료 → 육수 전면 금지, 재료 자체를 제외
                # 예전엔 '물'로 치환해 원래 무게(예: 멸치3g+다시마1.5g=물4.5g)를 그대로 남겼는데,
                # 이 무게는 애초에 "국물 만들 물의 양"이 아니라 육수용 재료의 무게일 뿐이라 재료표에
                # "물 4.5g"처럼 국 한 그릇 답지 않은 뜻모를 숫자로 남는 문제가 있었음(2026-07-24,
                # 사용자 지적). 원본 데이터 자체가 국물 물의 양을 재료·조리과정 어디에도 기록 안 해서
                # (예: '냄비에 물을 넣고'만 있고 g/ml 없음) 대신 채워 넣을 근거도 없음 → 그냥 제외.
                # 영양(E/protein/P/K/Na)엔 영향 없음(물은 전부 0이라 있으나 없으나 totals()는 동일).
                continue
            if rice_dish and is_japgok_grain(ing):               # 재료 대체: 잡곡 곡물 → 백미
                ing = WHITE_RICE_ING
            d = ing_nut.get(ing, {'E': None, 'protein': None, 'P': None, 'K': None, 'Na': None, 'group': ''})
            inst.append({'menu': disp, 'ing': ing, 'amt': a, 'orig': a,
                         'E': d['E'], 'protein': d['protein'], 'P': d['P'], 'K': d['K'], 'Na': d['Na'],
                         'group': d['group'], 'fixedK': False})
    # 같은 (메뉴, 재료) 병합: 여러 잡곡을 백미로 바꾸면 백미가 여러 조각 되므로 양을 합침
    # (오곡밥: 멥쌀45 + 수수→백미10 + 팥→백미10 + 기장→백미5 = 멥쌀,백미 70g 하나로).
    merged, order = {}, []
    for i in inst:
        key = (i['menu'], i['ing'])
        if key in merged:
            merged[key]['amt'] += i['amt']
            merged[key]['orig'] += i['orig']
        else:
            merged[key] = i
            order.append(key)
    return [merged[k] for k in order]


AMT_FLOOR = 0.5   # 재료 양은 원본의 50% 밑으로는 안 줄임 (한 반찬만 없애버리지 않게)
AMT_CEIL = 2.0    # 일반 재료는 원본의 2배까지 (실측 95%가 1.46배 이내. 3배는 0.3%만 걸러 무의미했음)
OIL_CEIL = 3.0    # 기름(유지류)만 예외 — 열량 보충용이고 소량이라 3배여도 1.5g→4.5g 수준
# 생쌀→밥 환산 약 2.5배. 밥 한 공기가 200~210g(임상 표준 식단도 쌀밥 210g)이므로
# 생쌀 80g이 한 공기. 하한 40g(밥 100g=반공기), 상한 110g(밥 275g=한공기 조금 넘게).
RICE_MIN_G = 40.0   # 주식(생쌀) 절대 하한 ~ 밥 100g
RICE_MAX_G = 110.0  # 주식(생쌀) 절대 상한 ~ 밥 275g


def _is_staple(i):
    return i['menu'] in RICE and '쌀' in i['ing']


def amt_floor_of(i):
    """이 재료를 줄일 수 있는 절대 하한 (주식은 g 기준, 나머지는 원본 비율)."""
    o = i.get('orig', i['amt'])
    return max(RICE_MIN_G, o * AMT_FLOOR) if _is_staple(i) else o * AMT_FLOOR


def amt_ceil_of(i):
    """이 재료를 늘릴 수 있는 상한 배율 (기름은 열량 보충용이라 조금 더 허용)."""
    return OIL_CEIL if i.get('group') == '유지류' else AMT_CEIL


def unrealistic_reason(inst):
    """양이 비현실적이면 이유 문자열, 정상이면 None.
    레버들이 각자 스케일을 곱해 배율이 누적되면 밥 1g·재료 11배 같은 결과가 나온다(실측).
    레버 안에서 강제로 자르면 레버끼리 싸워 영양이 오히려 깨지므로(71.6%→51%),
    자르지 않고 '후보를 고를 때' 이 검사로 거른다 — 앱은 어차피 후보를 48개 뽑는다."""
    for i in inst:
        o = i.get('orig')
        if not o or o <= 0:
            continue
        if is_seasoning(i):
            continue        # 조미료·젓갈은 나트륨 레버가 의도적으로 0.3배까지 줄임 = 정상(저염 조리)
        if _is_staple(i) and not (RICE_MIN_G <= i['amt'] <= RICE_MAX_G):
            return f"밥 양 {i['amt']:.0f}g (적정 {RICE_MIN_G:.0f}~{RICE_MAX_G:.0f}g)"
        r = i['amt'] / o
        if r < AMT_FLOOR - 0.05 or r > amt_ceil_of(i) + 0.05:
            return f"{i['ing'][:12]} 양이 원본의 {r:.1f}배"
    return None
DRIED = ('말린', '마른', '건조', '말랭', '반건')   # 건조식품 표식 (생것과 무게밀도 ~10배 차)

def reducible(i):
    return i['amt'] > i.get('orig', i['amt']) * AMT_FLOOR + 1e-6

def reduce_amt(i, factor):
    i['amt'] = max(i['amt'] * factor, i.get('orig', i['amt']) * AMT_FLOOR)


# 인 흡수율 보정 (임상영양사 지정, 2026-07-22): 식물성 유기인(피테이트)은 흡수율이 낮아 ×0.7,
# 가공식품 무기인(인산염 첨가)은 흡수율 ~100%로 유기인(~70%)보다 높아 ×1.5 — 0.7의 대칭 가중.
# P: 실제 섭취 인(원값) — 최종 통과/실패를 가르는 하드 기준(passes()·meal_bounds() 예산 이월 모두 이 값).
# Peff: 흡수 보정 인 — 하드 기준이 아니라, 대체재 후보를 고를 때 "더 흡수가 덜 되는 쪽"을 선호하는
# 최적화용 소프트 신호. Pmax 같은 임상 상한과 직접 비교해 pass/fail을 결정하는 데는 쓰지 않는다.
PLANT_PROTEIN_GROUPS = {'두류'}
PLANT_P_FACTOR = 0.7
PROCESSED_P_FACTOR = 1.5


def p_abs(P, group, ing=''):
    """흡수 보정 인(mg/100g). 가공식품(무기인) ×1.5, 식물성 단백질군(유기인) ×0.7, 그 외 원값.
    ※ is_processed_name은 이 함수보다 아래에 정의되지만 호출 시점엔 이미 로드돼 있어 문제없음."""
    if P is None:
        return None
    if is_processed_name(ing, group):
        return P * PROCESSED_P_FACTOR
    if group in PLANT_PROTEIN_GROUPS:
        return P * PLANT_P_FACTOR
    return P


def totals(inst):
    t = {'E': 0, 'protein': 0, 'P': 0, 'K': 0, 'Na': 0}
    for i in inst:
        for k in t:
            if i[k] is not None:
                t[k] += i['amt'] / 100 * i[k]
    # 흡수 보정 인 합계(최적화 소프트 신호 — 대체재 선호도 등 참고용. Pmax 하드 판정에는 t['P']를 쓴다).
    # 가공식품 1.5배·식물성 단백질군 0.7배 반영.
    t['Peff'] = sum(i['amt'] / 100 * p_abs(i['P'], i['group'], i['ing'])
                    for i in inst if i['P'] is not None)
    t['Na_season'] = season_na(inst)   # 첨가염(조미료)만 = 나트륨 판정용
    return t


def scale_menu(inst, menu, f, skip_seasoning=False):
    for i in inst:
        if i['menu'] == menu and not (skip_seasoning and is_seasoning(i)):
            i['amt'] *= f


def lever_kimchi(inst):
    # 김치 반찬 메뉴 → 저염김치 풀에서 로테이션 치환 (메뉴 통째 교체)
    ing_nut = NUT[0]; master = NUT[5]
    present = list(dict.fromkeys(i['menu'] for i in inst))
    for m in present:
        if m not in KIMCHI_SIDES:
            continue
        inst[:] = [i for i in inst if i['menu'] != m]
        choice = LOWNA_POOL[ROT[0] % len(LOWNA_POOL)]
        ROT[0] += 1
        for ing, amt in master[choice]:
            a = num(amt)
            if a is None:
                continue
            d = ing_nut.get(ing, {'E': None, 'protein': None, 'P': None, 'K': None, 'Na': None, 'group': ''})
            inst.append({'menu': choice, 'ing': ing, 'amt': a, 'orig': a,
                         'E': d['E'], 'protein': d['protein'], 'P': d['P'], 'K': d['K'], 'Na': d['Na'],
                         'group': d['group'], 'fixedK': False})


PROCESSED_KW = ('장아찌', '젓', '어묵', '맛살', '소시지', '햄', '베이컨', '미트볼',
                '인스턴트', '단무지', '훈제', '너겟', '완자', '스팸', '절임', '무침', '김치')
# 절임/무침/김치: DB상 '채소류'로 분류돼 is_processed 판정을 피해가는 고나트륨·발효 재료 방지
# (예: 오이지·염절임 Na1647mg, 무말랭이무침 Na1073mg, 쌈무 Na400mg, 김치 Na300~550mg — 대체
#  타겟으로 골라지면 칼륨 개선 명목으로 나트륨을 대량 주입하게 됨. 2026-07-24 자체 감사로 발견.
#  김치가 소스(원재료)로 뽑히는 경로는 _best_k_swap의 별도 startswith('김치') 가드로 막음 —
#  여기선 타겟(대체 후보)으로 뽑히는 반대 방향까지 막는다.)


def is_processed_name(name, group):
    """재료명·식품군으로 가공식품 판정 (대체 '후보'를 검사할 때 씀 — 아직 inst 항목이 아님)."""
    if group in ('조리가공식품류', '절임류'):
        return True
    return any(w in (name or '') for w in PROCESSED_KW)


# 대체 채택시 식품군 호환 판정 (2026-07-22 확정): 대체사전(apply_potassium/phosphorus_filter.py)이
# same_category(클러스터)로 만들어졌는데, 레버가 '정확히 같은 식품군'만 받아들여 새우완자→닭고기처럼
# 클러스터 안이지만 그룹명이 다른 항목이 통째로 죽어있었음(SUBS_P 26쌍 중 15쌍 dead 확인).
# ing_nut(식약청 원본)은 두류를 단백질/전분으로 안 나누지만, 그 구분은 사전을 만들 때(빌드타임) 이미
# 반영됐으므로(런타임에 나타나는 두류 후보=항상 두부류) 여기선 두류 전체를 단백질군에 포함해도 안전.
COMPATIBLE_GROUPS = [
    {'육류 및 그 제품', '어패류 및 그 제품', '난류', '조리가공식품류', '두류'},   # 단백질류
    {'채소류', '버섯류', '해조류', '절임류'},
    {'곡류 및 그 제품', '감자류 및 전분류', '빵 및 과자류'},
    {'조미료류'}, {'유지류'}, {'당류'}, {'과일류'}, {'우유 및 그 제품'}, {'견과류 및 종실류'},
]


def same_category(g1, g2):
    """두 식품군이 호환 클러스터 안에 있는지 (정확 일치보다 완화된 비교)."""
    if g1 is None or g2 is None:
        return False
    if g1 == g2:
        return True
    return any(g1 in grp and g2 in grp for grp in COMPATIBLE_GROUPS)


def is_processed(i):
    return is_processed_name(i['ing'], i['group'])


def lever_sodium(inst):
    # 조미료(첨가염)를 첨가 나트륨이 소금 1g(393mg) 되도록 적응적 축소. 하한 0.3×(맛 보존).
    # 기본 레시피가 너무 짜면(하한에서도 393 초과) 그대로 두고 → 판정에서 미달로 표시(앱 경고).
    # 축소 대상은 나트륨이 실질적인 조미료만(is_salty_seasoning) — 식초·후추·설탕처럼 나트륨이
    # 거의 없는 조미료류는 줄여도 목표치에 기여가 없어 원래 양 그대로 둔다.
    s = season_na(inst)
    if s <= SALT_MG:
        return
    f = max(SEASON_FLOOR, SALT_MG * 0.995 / s)   # 0.5% 여유: 경계 부동소수 통과 보장
    for i in inst:
        if is_salty_seasoning(i):                # 나트륨 있는 조미료 + 젓갈만 축소
            i['amt'] *= f


def _menu_na(inst):
    """메뉴별 총 나트륨 기여(mg). 첨가염(조미료)도 그 메뉴 몫으로 포함 — 메뉴를 줄이면 같이 줄기 때문."""
    m = defaultdict(float)
    for i in inst:
        m[i['menu']] += i['amt'] / 100 * (i['Na'] or 0)
    return m


def _menu_min_ratio(inst, menu, skip_seasoning=False):
    """이 메뉴가 원본 대비 현재 얼마까지 줄었나(가장 많이 줄어든 재료 기준). 1.0=원본, 0.5=하한.
    skip_seasoning=True면 조미료류 제외 — 조미료는 하한이 다름(SEASON_FLOOR=0.3, 나트륨 레버 소관).
    섞어서 계산하면 나트륨이 조미료를 먼저 깎았을 때 '이미 충분히 줄었다'고 오판해 다른 레버가
    멈춰버림(실측 버그: 시금치나물의 간장·소금이 0.3까지 줄자 칼륨 레버의 scale_menu가 통째로 먹통됨)."""
    rs = [i['amt'] / o for i in inst if i['menu'] == menu and not (skip_seasoning and is_seasoning(i))
          for o in (i.get('orig', i['amt']),) if o > 0]
    return min(rs) if rs else 1.0


def _is_main_ing(i, inst):
    """이 재료가 그 메뉴의 '메인'인가 → 메인이면 메뉴 통째로, 부재료면 그것만 줄인다.
    메인 판단: (1) 재료명이 메뉴명에 들어감(미역국의 미역, 오징어볶음의 오징어) 또는
              (2) 그 메뉴에서 물·조미료를 뺀 건더기 중 양이 가장 많은 재료(어묵볶음의 어묵)."""
    if is_identity_ingredient(i):
        return True
    solids = [x for x in inst if x['menu'] == i['menu'] and not is_seasoning(x)
              and x['ing'].split(',')[0].strip() not in ('물', '생수')]
    return bool(solids) and i is max(solids, key=lambda x: x['amt'])


def lever_sodium_extra(inst, namax):
    # 총 나트륨이 끼니 상한(namax)을 넘으면 나트륨 기여 큰 재료부터 줄인다.
    #   · 메인 재료(미역국의 미역)면 → 메뉴 통째로 비율 맞춰 축소(scale_menu): 미역·물·국간장 함께 줄어
    #     '작은 그릇'이 됨. 메인만 줄여 비율 깨지는 것 방지 + 그 메뉴 첨가염도 함께 감소.
    #   · 부재료(나물에 곁들인 소량 김 등)면 → 그 재료만 축소(reduce_amt).
    # 원본 50%까지만. 더 줄여도 못 맞추면(미역 없는 미역국은 성립 안 함) 앱이 '국물 남기기' 안내.
    snack_menus = {s[0] for s in SNACK_POOL}    # 간식은 열량 채우려 넣은 것 → 나트륨 축소 대상 아님
    for _ in range(30):
        if totals(inst)['Na'] <= namax:
            break
        best = None
        for i in inst:
            if not i['Na'] or i['amt'] <= 1 or is_seasoning(i) or i['menu'] in snack_menus:
                continue
            main = _is_main_ing(i, inst)
            can = (_menu_min_ratio(inst, i['menu']) > 0.5 + 1e-9) if main else reducible(i)
            if not can:
                continue
            na = i['amt'] / 100 * i['Na']
            if best is None or na > best[0]:
                best = (na, i, main)
        if best is None:
            break
        _, i, main = best
        if main:
            cur = _menu_min_ratio(inst, i['menu'])
            scale_menu(inst, i['menu'], max(0.8, 0.5 / cur))   # 메뉴 통째, 50% 하한
        else:
            reduce_amt(i, 0.8)                                  # 부재료만


# 임상 저칼륨 대체 (영양사 검수 대상). food-embedding subs가 놓치는 '저명한' 저칼륨 치환.
# 원재료 이름 접두 → 저칼륨 대체 재료(DB 정확명). 식품군이 달라도 조리역할이 비슷하면 허용
# (예: 감자=서류 → 무=채소류. subs는 '같은 군'만 되므로 감자·바나나 등은 대체가 없었음).
CLINICAL_K_SWAP = {
    '감자, 대지':   '무, 조선무, 생것',    # 서류 고칼륨 → 근채 (조림·국·볶음)
    '고구마':       '무, 조선무, 생것',
    '토란':         '무, 조선무, 생것',
    '시금치':       '양배추, 생것',        # 고칼륨 잎채소 → 저칼륨 잎채소 (나물·국)
    '근대':         '양배추, 생것',
    '아욱':         '양배추, 생것',
    '바나나':       '사과, 부사, 생것',    # 고칼륨 과일 → 저칼륨 과일
    '참외':         '사과, 부사, 생것',
    # '대두':'두부' 제거(2026-07-22, 사용자 지적으로 재검토): 원물 검은콩·서리태를 겨냥해 만들었으나
    # 실제 레시피에서 '대두' 접두에 걸리는 건 전부 콩가루뿐(콩가루배춧국 등 국물 걸쭉이 3건) —
    # 원물 콩 요리는 이 데이터에 아예 없음. 콩가루(가루)→두부(고체)는 제형이 안 맞아 조리법이 깨짐
    # (녹두→전분/밀가루 제외했던 것과 같은 종류의 문제). gain은 컸지만(82mg) 대가가 더 큼.
}

# 위 접두매칭이 잘못 잡는 '다른 부위' 제외어. 고구마줄기·토란대(줄기)는 나물 반찬 재료로,
# 고구마·토란(뿌리/알줄기, 국·조림용)과 조리역할이 전혀 다름(실측: 고구마순나물이 무로 잘못 교체됨).
STEM_EXCLUDE_WORDS = ('줄기', '순')


# 표시용 메뉴명 정리: 원본 데이터에 '돈까스소스'처럼 [메인+곁들임장]이 붙어 있어 소스만 먹는 것처럼
# 보인다(실제 재료엔 돼지고기·빵가루가 다 있음). 데이터는 그대로 두고 표시할 때만 '돈까스+소스'로 나눈다.
# (데이터 자체를 바꾸면 모델 토큰·레시피 키·식단 CSV가 서로 어긋날 위험이 큼)
SAUCE_SUFFIXES = ('타르타르소스', '머스타드소스', '머스터드소스', '양파크림소스',
                  '오리엔탈드레싱', '발사믹드레싱', '참깨드레싱', '키위드레싱',
                  '초고추장', '초간장', '달래장', '양념장', '쌈장', '드레싱', '소스')


# 표시 이름 변경 (데이터/키는 그대로, 화면에 보여줄 때만 다른 이름으로) — 이름은 안 바뀌지만
# 뜻은 두 갈래: (1)뼈육수 메뉴 → 육수 재료가 맹물로 바뀌면 사실상 '소고기국류'라 이름도 맞춰 표시
# (뼈육수는 못 먹고 인이 높아 금지 → 남는 건 소고기·건더기뿐. 꼬리곰탕은 고기가 없어 별도 제외).
# (2)'성인' 계열(2026-07-22): 대상자가 전부 성인이라 '성인'이 붙은 큰 분량판을 표준으로 쓰되
# 이름에서 '성인'을 뗌 — 원본(소아용 추정, 작은 분량)은 app_core_FOOK.py의 BLOCK_MENUS로 생성
# 자체를 막아 표시이름 충돌을 방지(둘 다 살아있으면 같은 이름 아래 다른 레시피가 나올 위험).
MENU_DISPLAY_RENAME = {
    '사골우거지국': '우거지된장국',
    '설렁탕': '소고기맑은국',      # 소고기 맑은국 대표 1개만 유지 (사골곰국·사골곰탕은 중복이라 제외)
    '우사골파국': '소고기달걀파국',
    '채소달걀말이성인': '채소달걀말이',
    '토마토달걀볶음성인': '토마토달걀볶음',
    # 참깨·들깨·견과류 제거 후 이름-내용 불일치 정리(2026-07-24). 병합은 안 되고(레시피가 실제로
    # 다름) 이름에만 참깨/들깨/견과류 단어가 남은 경우를 실제 재료 차이로 구분해 재명명.
    '연근아몬드조림': '연근새콤조림',       # 사과식초 기반이라 새콤한 맛이 특징(연근조림과는 별개 유지)
    '무채들깨국': '무채국',               # 소금 기본 간 (들깨무채국=간장무채국과 구분)
    '들깨무채국': '간장무채국',            # 간장 기본 간, 마늘 포함
    '버섯들깨국': '모둠버섯국',            # 느타리·표고·새송이·팽이 5종+무+양파, 큰 국
    '들깨버섯국': '버섯국',               # 팽이·표고 2종만, 단출한 국
}


def display_menu_name(menu):
    """'돈까스소스' -> '돈까스+소스', '설렁탕' -> '소고기맑은국'(뼈육수 제거). 해당 없으면 그대로."""
    if menu in MENU_DISPLAY_RENAME:
        return MENU_DISPLAY_RENAME[menu]
    for suf in SAUCE_SUFFIXES:          # 긴 접미사부터 매칭
        if menu.endswith(suf) and len(menu) > len(suf):
            return f'{menu[:-len(suf)]}+{suf}'
    return menu


def is_identity_ingredient(i):
    """재료의 핵심어가 메뉴 이름에 들어있는가 (오징어볶음의 '오징어').
    앵커(유저 지정메뉴)면 이 재료를 바꾸지 않고, 아니면 바꾸되 메뉴 이름도 함께 갱신한다."""
    core = i['ing'].split(',')[0].strip()
    return bool(core) and core in i['menu']


def rename_menu_for_swap(inst, menu, old_ing, new_ing):
    """메뉴 이름에 든 재료를 교체했으면 이름도 갱신 (오징어볶음 + 오징어→고등어 = 고등어볶음).
    안 그러면 '오징어볶음인데 고등어가 든' 이름-내용 불일치가 생김(실측 143종).
    원래 이름은 orig_menu에 남겨 조리법 조회에 씀."""
    old_core = old_ing.split(',')[0].strip()
    new_core = new_ing.split(',')[0].strip()
    if not old_core or not new_core or old_core == new_core or old_core not in menu:
        return menu
    # 재료명 뒤에 부위어(순/줄기)가 메뉴이름에 바로 이어지면 그것까지 포함해서 치환.
    # 안 그러면 '고구마순나물'→'무순나물'처럼 '무순'(무의 새싹, 다른 실존 채소)과 이름이 겹침
    # (2026-07-22, 실측으로 발견·수정). 토란대나물→무나물처럼 원래도 자연스러운 경우는 영향 없음.
    idx = menu.find(old_core)
    tail = menu[idx + len(old_core):]
    replace_core = old_core
    for w in STEM_EXCLUDE_WORDS:
        if tail.startswith(w):
            replace_core = old_core + w
            break
    new_menu = menu.replace(replace_core, new_core)
    for i in inst:
        if i['menu'] == menu:
            i.setdefault('orig_menu', menu)      # 조리법은 원래 메뉴 기준으로 찾아야 함
            i['menu'] = new_menu
    return new_menu


K_SWAP_MIN_GAIN = 5.0   # 칼륨 절감이 이보다 작으면 교체하지 않음.
# 실측: 설탕(K=1)→물엿(K=0) 교체가 46회 일어났는데 실제 절감은 0.06mg — 레시피만 바뀌고 효과가 없다.

# 인도 동일 원칙(2026-07-22, K와 대칭). 기존엔 문턱이 아예 없어(g>0이면 무조건 통과) 구조적 위험이었음
# — 실측 최소사례(부대찌개 소시지7.5g→닭고기)가 gain 4.58mg로 이 문턱(5.0) 바로 아래라 딱 걸러짐,
# 나머지 실사용(대부분 25g+)은 전혀 영향 없음(K SWAP_MIN_GAIN과 마찬가지로 향후 사전 확장 대비 안전장치).
P_SWAP_MIN_GAIN = 5.0

# 향신채(마늘·생강)는 소량으로 맛의 베이스를 결정하므로 일반 재료보다 높은 절감기준 적용(2026-07-22).
# 실측: 마늘 사용량 537건 중 82%가 1.71g 미만(=5mg 기준 미만, 원래도 교체 안 됨)이지만 나머지 18%
# 중 2~3g대(찜·조림·볶음 양념 베이스)에서도 실제 교체가 발동(단호박해물찜 등, 15건/365일).
# '마늘,'/'생강,' 콤마매칭 — 마늘종(꽃줄기, 채소 주재료 30~105g)은 이름이 달라 자동 제외됨.
# 20mg 기준이면: 양념 베이스(2~3g, gain 6~9mg)는 걸러지고, 통마늘 요리(12g, gain 35mg, 제육통마늘볶음
# 등 마늘 자체가 주재료급)는 그대로 살아남음 — 그램 하드코딩 없이 자연 분리됨.
AROMATIC_KEYS = ('마늘,', '생강,')
K_SWAP_MIN_GAIN_AROMATIC = 20.0


def _k_swap_min_gain(ing):
    return K_SWAP_MIN_GAIN_AROMATIC if ing.startswith(AROMATIC_KEYS) else K_SWAP_MIN_GAIN


def menu_has_ingredient(pool, menu, ing_name, exclude=None):
    """같은 메뉴 안에 이미 같은 기본식품 재료가 있는지 (자기자신 제외).
    새우완자·미더덕이 같은 메뉴에서 각자 가자미로 바뀌어 '가자미, 가자미' 중복이 되는 것을 막는다."""
    base = ing_name.split(',')[0].strip()
    return any(x is not exclude and x['menu'] == menu and x['ing'].split(',')[0].strip() == base
               for x in pool)


def is_sole_solid_ingredient(pool, menu, exclude):
    """이 재료가 그 메뉴에서 사실상 유일한 고형재료인가(조미료·기름·당류 제외한 다른 게 없음).
    조림·볶음·국처럼 여러 채소가 섞인 요리는 하나를 바꿔도 자연스럽지만(감자조림→무조림 등 실존 요리),
    '삶은감자'·'찐고구마'·'고구마맛탕'처럼 그 재료 하나가 메뉴 전부면 대체 시 '삶은무'·'찐무'·'무맛탕'처럼
    존재하지 않는 요리가 됨 — 이런 메뉴는 대체(CLINICAL_K_SWAP) 대상에서 제외, 양조절로만 대응.
    (2026-07-22, 실측: 감자/고구마 90여개 메뉴 중 이 패턴은 단 4개뿐이었음)"""
    return not any(x is not exclude and x['menu'] == menu
                   and x['group'] not in ('조미료류', '유지류', '당류')
                   for x in pool)


def _best_k_swap(pool, protect_identity=False):
    """pool 안에서 칼륨 절감이 가장 큰 교체 후보 1개. 없으면 None.
    임상 저칼륨 대체(CLINICAL_K_SWAP)를 먼저 시도(식품군 우회), 없으면 subs(같은 군).
    protect_identity=True(앵커)면 메뉴 이름에 든 재료는 건드리지 않는다."""
    ing_nut, _, ing2kw, kw_rep, subs, _ = NUT
    best = None
    for i in pool:
        if i['fixedK'] or i['K'] is None:
            continue
        if i['group'] == '조미료류':          # 조미료는 교체 안 함(맛 보존), 양만 조절
            continue
        if any(d in i['ing'] for d in DRIED):  # 건조식품은 교체 안 함(무게밀도 왜곡=가짜감소)
            continue
        if i['ing'].startswith('김치'):        # 김치는 발효·양념 식품이라 생채소로 교체 불가
            continue                            # (DB상 식품군은 '채소류'라 same_category를 통과해버림)
        if protect_identity and is_identity_ingredient(i):   # 앵커의 정체성 재료는 보존
            continue
        # (1) 임상 저칼륨 대체 우선 (명시 매핑, 식품군 무시)
        # 고구마줄기·토란대(줄기 부위)도 허용(2026-07-22, 재검토): 나물↔나물 대체 자체는 문제없음
        # (시금치→양배추와 같은 결). 단 이름표시는 rename_menu_for_swap에서 부위어까지 포함해 치환
        # (안 그러면 '고구마순나물'→'무순나물'처럼 무순=무의 새싹이라는 다른 실존채소와 이름이 겹침).
        # 단, 그 재료가 메뉴의 유일한 고형재료면(삶은감자·찐고구마·고구마맛탕·상추쌈쌈장 등) 대체 제외 —
        # '삶은무'·'찐무'·'무맛탕'·'오이쌈쌈장'은 존재하지 않는 요리라 대체보다 양조절이 맞음(2026-07-22).
        # (1)CLINICAL_K_SWAP뿐 아니라 (2)subs(자동생성 사전)에도 동일 적용 — 실측: 상추쌈쌈장·미숫가루가
        # subs 경로로 실제 위험하게 발동한 것 확인해 확장.
        cand = None
        sole = is_sole_solid_ingredient(pool, i['menu'], exclude=i)
        if not sole:
            for key, rep in CLINICAL_K_SWAP.items():
                if i['ing'].startswith(key):
                    nd = ing_nut.get(rep)
                    if nd and nd['K'] is not None and nd['K'] < i['K'] and \
                            not is_processed_name(rep, nd['group']) and \
                            (nd['P'] is None or i['P'] is None or nd['P'] <= i['P'] * 1.2) and \
                            not menu_has_ingredient(pool, i['menu'], rep, exclude=i):   # 같은 메뉴 중복 방지
                        cand = (i['amt'] / 100 * (i['K'] - nd['K']), i, rep, nd)
                    break
        # (2) 없으면 subs (같은 식품군만, 역할 보존)
        if cand is None and not sole:
            for sub in subs.get(ing2kw.get(i['ing']), []):
                rep = kw_rep.get(sub)
                nd = ing_nut.get(rep) if rep else None
                if not nd or nd['K'] is None or not same_category(nd['group'], i['group']):
                    continue        # 호환 클러스터만(역할 보존) — 정확일치보다 완화
                if rep.split(',')[0].strip() == i['ing'].split(',')[0].strip():
                    continue        # 같은 기본식품끼리 교체 금지(자기자신)
                if is_processed_name(rep, nd['group']) and not is_processed(i):
                    continue        # 자연식품 → 가공식품 교체 금지 (게→게맛살: 인산염 첨가 위험)
                if menu_has_ingredient(pool, i['menu'], rep, exclude=i):
                    continue        # 같은 메뉴에 이미 같은 재료 있음 (중복 방지)
                if nd['K'] < i['K'] and (nd['P'] is None or i['P'] is None or nd['P'] <= i['P'] * 1.2):
                    cand = (i['amt'] / 100 * (i['K'] - nd['K']), i, rep, nd)
                    break
        if cand and cand[0] >= _k_swap_min_gain(i['ing']) and (best is None or cand[0] > best[0]):
            best = cand     # 절감이 미미한 교체는 하지 않음(레시피만 바뀜) — 향신채는 기준 더 높음
    return best


def lever_potassium(inst, kmax, anchor=None):
    """칼륨 상한 초과시 재료 교체. 유저 지정메뉴(anchor)는 최후수단으로만 건드림
    (금지가 아니라 후순위 -> 통과율은 그대로, 앵커는 가능할 때만 보존).
    순서: (1)비앵커 교체 (2)비앵커 양 감소(2026-07-22, 인 레버와 대칭 — 임상 안전성 위해 비앵커도
    교체 실패시 포기 대신 양조절 폴백 추가) (3)앵커 교체 (4)그래도 없으면 앵커 메뉴 전체를 비율
    맞춰 통째로 축소(scale_menu, 나트륨 레버와 동일 원칙) — 정체성 재료만 줄이면 마늘·참기름 등
    나머지 양념 비율이 깨지므로 메뉴 전체를 같은 비율로. 하한 50%(2026-07-22, 사용자 지적).
    앵커는 교체·양감소 둘 다 비앵커가 완전히 소진된 뒤에만 건드림(최후수단 원칙 강화)."""
    for _ in range(20):
        if totals(inst)['K'] < kmax:          # 칼륨 '미만' 기준
            break
        non = [i for i in inst if i['menu'] != anchor]
        best = _best_k_swap(non)              # (1) 비앵커 교체
        if best is None:
            # (2) 비앵커 양 감소 — 교체 후보가 소진돼도 앵커를 건드리기 전에 먼저 시도(인 레버와 동일 원칙)
            cand = [i for i in non if i['K'] and i['amt'] > 1 and reducible(i)]
            if cand:
                reduce_amt(max(cand, key=lambda x: x['amt'] / 100 * x['K']), 0.7)
                continue
            # (3) 그래도 없으면 앵커는 최후수단으로 교체
            best = _best_k_swap([i for i in inst if i['menu'] == anchor], protect_identity=True)
        if best is not None:
            _, i, rep, nd = best
            SWAP_LOG.append((i['menu'], i['ing'], rep, 'K', i['K'], nd['K']))
            old_ing, old_menu = i['ing'], i['menu']
            i['ing'] = rep
            for k in ('E', 'protein', 'P', 'K', 'Na', 'group'):
                i[k] = nd[k]
            i['fixedK'] = True
            rename_menu_for_swap(inst, old_menu, old_ing, rep)   # 이름에 든 재료였으면 메뉴명도 갱신
            continue
        # (4) 교체 후보가 아예 없으면(정체성 보호+대체소진) 앵커 메뉴에서 '조미료 제외' 재료를
        # 비율 맞춰 통째 축소. 조미료는 나트륨 레버 소관(SEASON_FLOOR=0.3)이라 건드리지 않음 —
        # 섞으면 나트륨이 먼저 깎은 조미료 때문에 _menu_min_ratio가 낮게 잡혀 이 단계가 먹통됨
        # (실측 버그: 시금치나물 간장·소금이 0.3까지 줄자 이 단계가 통째로 스킵돼 시금치가 전혀 안 줄음).
        has_ident = any(i['menu'] == anchor and i['K'] and is_identity_ingredient(i) for i in inst)
        cur = _menu_min_ratio(inst, anchor, skip_seasoning=True)
        if has_ident and cur > 0.5 + 1e-9:
            scale_menu(inst, anchor, max(0.8, 0.5 / cur), skip_seasoning=True)
            continue
        break


def lever_phosphorus(inst, pmax, anchor=None, plo=0):
    """인 상한 초과 조정. 유저 지정메뉴(anchor)의 재료는 교체 안 함(정체성 보존).
    순서: (1)비앵커 저인 대체[가공→생것 / subs 저인, 단백질 75%↑ 유지=커플링대응]
          (2)비앵커 양 감소  (3)앵커 양 감소(단백질 하한 plo 지킴)  (4)더 못 줄이면 실패.
    반환: 상한 충족 성공 여부(False면 앵커 지키느라 인 초과 → 앱이 유저에 알림).

    P: 실제 섭취 인(raw P) = 하드 기준 — app_core_FOOK.passes()의 최종 통과 판정과 동일 기준.
    Peff: 흡수보정 인 = 최적화용 소프트 신호(대체재가 여럿일 때 그중 더 낮은 쪽을 고르는 선호도).
    [2026-08-11] 진입/수렴/랭킹 기준을 원래 Peff로 뒀던 버그를 raw P로 통일했다 — Peff<pmax로
    수렴한 뒤 이후 단백질·나트륨·열량 레버가 되흔들면(phosphorus↔protein 핑퐁) raw P는 상한을
    넘겼는데도 이 레버 자신은 "통과"로 착각하고 멈추는 사례가 있었다(passes()는 raw P만 보므로
    실제 서비스 판정과 어긋남). lever_phosphorus_rawP(두부·콩류 앵커 전용 경로)는 애초에 이 4곳을
    raw P로 뒀던 게 정답이었고, 이 함수(일반 경로)도 이제 동일 기준을 쓴다."""
    ing_nut, base_fresh, ing2kw, kw_rep, subs, _ = NUT
    for _ in range(25):
        if totals(inst)['P'] < pmax:          # raw P '미만' 기준 — passes()와 동일한 하드 기준
            return True
        non = [i for i in inst if i['menu'] != anchor]

        # (1) 비앵커 저인 대체
        best = None
        for i in non:
            if i['P'] is None:
                continue
            if i['group'] == '조미료류':                     # 조미료는 교체 안 함(맛 보존)
                continue
            # 건조식품(미역·다시마·멸치·당면 등)은 생것과 무게밀도가 ~10배 달라
            # 같은 g으로 교체하면 실제 먹는 양이 급감 = 가짜 감소. → 교체 안 하고 양감소로만 대응.
            if any(d in i['ing'] for d in DRIED):
                continue
            # (구) '가공형태 → 같은 기본식품 생것' 경로 제거: base_fresh는 첫 단어(기본식품명)로만
            # 후보를 찾아 항상 같은 식품의 다른 품종/부위를 골랐다. 그 결과 새우살→새우껍질(P197→3),
            # 들깨→들깨싹, 건포도→생포도처럼 먹을 수 없거나 요리가 달라지는 '가짜 인 감소'가 나왔다.
            # 인 대체는 '진짜 다른 식품'(인 전용 SUBS_P)과 양 감소로만 처리한다.
            # ※ 칼륨 subs 재활용 금지 — SUBS_P는 P/단백질 비율 기반(현미→쌀, 가공→자연 등).
            if is_sole_solid_ingredient(non, i['menu'], exclude=i):
                continue   # 유일한 고형재료면 대체 대신 양조절로(K레버와 동일 원칙, 2026-07-22)
            # 채택판정은 raw P(하드 기준, passes()와 동일) — 이 재료(i)의 대체재 후보 중 raw P를
            # 실제로 줄이면서 커플링(단백질 75%↑) 조건을 만족하는 것들을 전부 모은 뒤, raw P
            # 감소량이 가장 크고, 동률이면 흡수보정 인(Peff)이 더 낮은 후보를 고른다 — Peff는
            # "이미 raw P 조건을 만족하는 후보들 사이의 선호도"로만 쓰인다(2026-08-11 추가).
            sub_best = None   # (g, rep, nd, effP_nd)
            for sub in SUBS_P.get(ing2kw.get(i['ing']), []):   # 저인비율 대체재 (같은군 + 단백질 유지)
                rep = kw_rep.get(sub); nd = ing_nut.get(rep) if rep else None
                if not nd or nd['P'] is None or not same_category(nd['group'], i['group']):
                    continue                                 # 호환 클러스터만(역할 보존) — 정확일치보다 완화
                if rep.split(',')[0].strip() == i['ing'].split(',')[0].strip():
                    continue                                 # 같은 기본식품끼리 교체 금지(밀→밀 등 자기자신)
                if is_processed_name(rep, nd['group']) and not is_processed(i):
                    continue                                 # 자연식품 → 가공식품 교체 금지
                if menu_has_ingredient(non, i['menu'], rep, exclude=i):
                    continue                                 # 같은 메뉴에 이미 같은 재료 있음 (중복 방지)
                if nd['P'] < i['P']:                          # raw P 기준 감소 여부(하드 기준)
                    ip, npr = (i['protein'] or 0), (nd['protein'] or 0)
                    if ip == 0 or npr >= ip * 0.75:          # 단백질 75%↑ 유지 = 커플링 대응
                        g = i['amt'] / 100 * (i['P'] - nd['P'])   # raw P 감소량 랭킹(하드 기준)
                        if g >= P_SWAP_MIN_GAIN:
                            effP_nd = p_abs(nd['P'], nd['group'], rep)   # Peff = 소프트 선호도
                            if (sub_best is None or g > sub_best[0]
                                    or (g == sub_best[0] and effP_nd < sub_best[3])):
                                sub_best = (g, rep, nd, effP_nd)
            if sub_best:
                g, rep, nd, _ = sub_best
                if best is None or g > best[0]:
                    best = (g, i, rep, nd)
        if best:
            _, i, rep, nd = best
            SWAP_LOG.append((i['menu'], i['ing'], rep, 'P', i['P'], nd['P']))
            old_ing, old_menu = i['ing'], i['menu']
            i['ing'] = rep
            for k in ('E', 'protein', 'P', 'K', 'Na', 'group'):
                i[k] = nd[k]
            rename_menu_for_swap(inst, old_menu, old_ing, rep)   # 이름에 든 재료였으면 메뉴명도 갱신
            continue

        # (2) 비앵커 양 감소 (raw P 기여 큰 재료) — 원본 50% 하한
        # factor 0.8→0.7(2026-07-22 조율): 가공육 1.5가중으로 드러난 위반을 양감소가 더 빨리 흡수.
        # 스윕 결과(365일): 0.8=인305 / 0.7=인312·전부충족260 / 0.5(플로어 직행)=인314지만 감소가 급격해 배제.
        cand = [i for i in non if i['P'] and i['amt'] > 1 and reducible(i)]
        if cand:
            reduce_amt(max(cand, key=lambda x: x['amt'] / 100 * x['P']), 0.7)
            continue

        # (3) 앵커(유저 메뉴) 양 감소 — 단백질 하한 + 원본 50% 하한
        anc = [i for i in inst if i['menu'] == anchor and i['amt'] > 1 and reducible(i)]
        if anc and totals(inst)['protein'] > plo:
            for i in anc:
                reduce_amt(i, 0.85)
            continue

        return False   # 앵커 지키느라 상한 못 맞춤 → 유저에게 알림
    return totals(inst)['P'] < pmax


def lever_phosphorus_rawP(inst, pmax, anchor=None, plo=0):
    """lever_phosphorus의 raw P 기준 버전(2026-07-27, 두부·콩류 앵커 전용 조건부 경로)이었다.
    [2026-08-11] lever_phosphorus 쪽의 판정 기준을 Peff→raw P로 통일하면서 두 함수의 로직이
    완전히 같아져(둘 다 raw P 하드 기준 + Peff 소프트 선호도), 중복을 없애고 lever_phosphorus에
    위임한다. 두부·콩류 전용 경로라는 호출부(adjust()의 _plant_protein_path_needed 분기,
    lever_protein_capped/lever_calorie_capped와 짝을 이루는 이름)는 그대로 유지 — 함수명·시그니처
    변경 없이 내부 구현만 통합했다."""
    return lever_phosphorus(inst, pmax, anchor=anchor, plo=plo)


def lever_protein(inst, lo, hi, anchor=None):
    """단백질찬(메뉴 통째)을 같은 비율로 스케일 → 맛/비율 유지.
    앵커(유저메뉴)는 후순위: 비앵커 메뉴 하나만 조절해서 목표에 닿을 수 있으면 그걸 쓰고,
    아무도 단독으로 못 맞출 때만 기여 최대 메뉴(=앵커일 수 있음)를 클램프해서 스케일."""
    t = totals(inst)['protein']
    if lo <= t <= hi:
        return
    pm = defaultdict(float)
    for i in inst:
        if i['protein']:
            pm[i['menu']] += i['amt'] / 100 * i['protein']
    if not pm:
        return
    target = (lo + hi) / 2

    # (1) 비앵커 중 단독으로 목표 달성 가능한 메뉴 (기여 큰 순)
    for m in sorted([x for x in pm if x != anchor], key=pm.get, reverse=True):
        cur = pm[m]
        new = target - (t - cur)
        if cur > 0 and new > 0 and 0.3 <= new / cur <= 2.0:
            scale_menu(inst, m, new / cur)
            return

    # (2) 아무도 단독으로 못 맞춤 → 기여 최대 메뉴, 클램프
    top = max(pm, key=pm.get)
    cur_top = pm[top]
    new_top = target - (t - cur_top)
    if cur_top > 0 and new_top > 0:
        scale_menu(inst, top, max(0.3, min(new_top / cur_top, 2.0)))


PROTEIN_CALORIE_CAP_FRAC = 0.90   # 두부·콩류 앵커 전용 raw P 예산 cap 배율(2026-07-27 확정,
# protein_phosphorus_final_compare/cross_anchor 실험에서 검증된 값. 0.80/1.00 스윕은 세 앵커
# 어디서도 필요성이 확인되지 않아 미실시 — frac=0.90 그대로 채택).


def _cap_scale_menu_rawP(inst, menu, ratio, pmax):
    """scale_menu의 raw P 예산판 — 비율>1(증량)일 때만 개입, 남은 raw P 예산의
    PROTEIN_CALORIE_CAP_FRAC(90%)까지만 늘어나게 배율을 제한한다. 감량(비율<=1)은 원본
    scale_menu와 완전히 동일(무개입). lever_protein_capped/lever_calorie_capped 전용
    보조함수(2026-07-27, 두부·콩류 앵커 전용 조건부 경로)."""
    if ratio <= 1.0 + 1e-9:
        scale_menu(inst, menu, ratio)
        return
    current_raw_P = totals(inst)['P']
    headroom = pmax - current_raw_P
    allowed_increase = max(0.0, headroom * PROTEIN_CALORIE_CAP_FRAC)
    menu_items = [i for i in inst if i['menu'] == menu]
    P_menu = sum(i['amt'] / 100 * i['P'] for i in menu_items if i['P'] is not None)
    Amt_menu = sum(i['amt'] for i in menu_items)
    if Amt_menu <= 0:
        scale_menu(inst, menu, ratio)
        return
    requested_grams = Amt_menu * (ratio - 1.0)
    density = P_menu / Amt_menu
    allowed_grams = float('inf') if density <= 1e-9 else allowed_increase / density
    applied_grams = max(0.0, min(requested_grams, allowed_grams))
    final_ratio = 1.0 + applied_grams / Amt_menu
    scale_menu(inst, menu, final_ratio)


def lever_protein_capped(inst, lo, hi, anchor=None, pmax=None):
    """lever_protein의 raw P 예산 cap 버전(2026-07-27, 두부·콩류 앵커 전용 조건부 경로).
    메뉴 선택 로직은 lever_protein과 완전히 동일 — 증량(scale_menu 비율>1)에만
    _cap_scale_menu_rawP로 남은 raw P 예산 90% 상한을 적용. 감량은 무변경.
    채택 근거는 lever_phosphorus_rawP 주석 참고(교차앵커 실험 결과)."""
    t = totals(inst)['protein']
    if lo <= t <= hi:
        return
    pm = defaultdict(float)
    for i in inst:
        if i['protein']:
            pm[i['menu']] += i['amt'] / 100 * i['protein']
    if not pm:
        return
    target = (lo + hi) / 2

    for m in sorted([x for x in pm if x != anchor], key=pm.get, reverse=True):
        cur = pm[m]
        new = target - (t - cur)
        if cur > 0 and new > 0 and 0.3 <= new / cur <= 2.0:
            _cap_scale_menu_rawP(inst, m, new / cur, pmax)
            return

    top = max(pm, key=pm.get)
    cur_top = pm[top]
    new_top = target - (t - cur_top)
    if cur_top > 0 and new_top > 0:
        ratio = max(0.3, min(new_top / cur_top, 2.0))
        _cap_scale_menu_rawP(inst, top, ratio, pmax)


RICE_FLOOR = 0.4   # 밥계열을 원래 양의 40% 밑으로는 안 줄임 (한 끼가 밥 한술 되는 것 방지)


# 데이터에 없지만 투석에 좋은 저칼륨 간식 추가 (load_all에서 master+MENU_CLASS에 등록 → 풀에 편입).
# 재료는 식약청 DB명으로 매핑. 양은 1인분 기준(가래떡 한 줄·식빵 1장·절편 2~3개·찹쌀떡 1개).
EXTRA_SNACK_RECIPES = {
    '가래떡': [('멥쌀떡, 가래떡', 80)],
    '절편': [('멥쌀떡, 절편', 50)],
    '식빵': [('빵, 식빵', 35)],
    '찹쌀떡': [('찹쌀떡', 60)],
}

# 저칼륨 과일 간식 — 임상 권장 1회 섭취량 기준(사용자 제공, 2026-07-23). 기존 raw데이터의
# 단감(100g)·딸기(150g)는 일반 인구 기준 분량이라 여기서 안전 분량으로 덮어씀. 나머지는 신규 등록.
EXTRA_FRUIT_SNACK_RECIPES = {
    '단감':      [('감, 단감, 생것', 75)],
    '연시':      [('감, 떫은감, 연시, 생것', 75)],
    '딸기':      [('딸기, 개량종, 생것', 80)],
    '포도':      [('포도, 캠벨얼리, 생것', 80)],
    '블루베리':    [('블루베리, 생것', 55)],
    '자두':      [('자두, 생것', 50)],
    '파인애플':    [('파인애플, 생것', 55)],
    '금귤':      [('금귤, 생것', 50)],
    '통조림백도':   [('복숭아, 백도, 통조림, 고형물', 65)],
    '통조림황도':   [('복숭아, 황도, 통조림, 고형물', 65)],
    '통조림프루트칵테일': [('프루트칵테일, 통조림', 65)],
}

# 원본 raw데이터 자체가 틀린 경우의 수정 override(raw 파일은 안 건드리고 병합 결과만 덮어씀).
# 너비아니구이: 9월 원본 raw의 정통 레시피 채택(한우 등심, 배즙양념, 얇게 썰어 두드려 굽기).
RECIPE_FIX_OVERRIDE = {
    '너비아니구이': [
        ('소고기, 한우, 등심, 생것', 52.5),   # 원본은 '1++등급'이나 DB에 등급별 세분류가 없어 일반값으로 매핑
        ('배즙', 15),   # 식약청_영양성분10.4에 별도 항목 존재(E38·단백0.23·K63·P12·Na2, 생배보다 낮음)
        ('파, 생것', 12),
        ('간장, 재래', 7.5),
        ('마늘, 생것', 6),
        ('설탕, 백설탕', 3.75),
        ('콩기름', 1.5),
        ('참기름', 0.75),
        ('참깨, 흰색, 볶은것', 0.3),
        ('후추, 검은색, 가루', 0.15),
    ],
}

# 열량 보충용 간식 (저칼륨·저인). load_all()에서 Class가 '~(간식)'인 메뉴 중 조건 맞는 것만 채움.
SNACK_POOL = []          # [(메뉴명, 열량, 칼륨, 인)] 열량 오름차순
SNACK_MAX_K = 150.0      # 간식 하나가 칼륨을 150mg 넘게 올리면 제외
SNACK_MAX_P = 120.0      # 인 120mg 넘으면 제외
SNACK_MIN_E = 35.0       # 이보다 적으면 보충 효과 없음(차 배제). 저칼륨 과일(귤·사과 39~53)은 포함
# 과일(간식)은 목적이 열량 보충이 아니라 수분/비타민 보충·기호 충족이므로 SNACK_MIN_E 예외
# (임상 안전 서빙량 자체가 작아 열량도 자연히 작음 — 예: 자두 50g=13kcal). 차/음료는 이름으로 이미 배제.
FRUIT_SNACK_CLASS = '과일(간식)'
# 투석 간식 임상 지침에 따른 배제 항목:
#   음료·차(열량 보충 안 됨) / 유제품·견과·초콜릿·탄산(인 높음) / 잡곡(인 높음) / 팥·콩 고물(칼륨·인)
SNACK_EXCLUDE_KW = ('차', '음료', '주스', '식혜', '수정과',
                    '우유', '요구르트', '두유', '치즈', '요플레',
                    '견과', '땅콩', '호두', '아몬드', '잣', '초콜릿', '초코',
                    '통밀', '호밀', '현미', '잡곡', '미숫', '팥', '콩')
# 농진청 표준식품성분표 저/중/고칼륨 과일 분류(사용자 제공, 2026-07-23) 기준, 간식은 저칼륨군만
# 추천하기로 함 → 필터를 통과해도 중칼륨군인 것은 별도로 배제. 딸기는 5알/80g 이내면 저칼륨이라는
# 사용자 지침대로 이미 그 서빙량으로 등록돼 있어 제외 대상 아님.
MID_K_FRUIT_EXCLUDE = {'귤', '감귤', '배', '수박'}
# 저칼륨군이어도 국내에서 실제로 흔히 먹는 과일이 아니면 배제(사용자 판단, 2026-07-23).
FRUIT_SNACK_LOW_AVAIL_EXCLUDE = {'금귤'}


def _build_snack_pool():
    """간식 후보 = 빵·떡·저칼륨 과일 위주 (투석 간식 지침). 저칼륨·저인·열량 있는 것만."""
    SNACK_POOL.clear()
    master = NUT[5]
    for menu, cls in MENU_CLASS.items():
        if '간식' not in str(cls) or menu not in master:
            continue
        if any(w in menu for w in SNACK_EXCLUDE_KW):
            continue
        if menu in MID_K_FRUIT_EXCLUDE or menu in FRUIT_SNACK_LOW_AVAIL_EXCLUDE:
            continue
        t = totals(expand([menu]))
        min_e = 0 if cls == FRUIT_SNACK_CLASS else SNACK_MIN_E
        if t['E'] < min_e or t['K'] > SNACK_MAX_K or t['P'] > SNACK_MAX_P:
            continue
        SNACK_POOL.append((menu, t['E'], t['K'], t['P']))
    SNACK_POOL.sort(key=lambda x: x[1])


_SNACK_BUILT = [False]


def add_snack(inst, need_kcal, kmax=None, pmax=None):
    """부족 열량에 가장 가까운 간식 1개를 한 끼에 추가. 성공하면 True.
    kmax/pmax를 주면 그 끼니의 현재 칼륨/인에 간식을 더해도 한도를 넘지 않는 후보만 고른다
    (마지막 레버 패스에서 호출되므로 이후 검증 기회가 없음 — 여기서 직접 걸러야 함, 2026-07-23)."""
    if not _SNACK_BUILT[0]:                   # NUT 로딩 후 첫 호출 때 후보 구성
        _build_snack_pool(); _SNACK_BUILT[0] = True
    if not SNACK_POOL or need_kcal < 10:      # 자두(13kcal) 등 저열량 과일 간식도 선택되게 문턱 하향
        return False                          # (마지막 레버 패스에서만 호출되므로 되먹임 없음, 5kcal 미만은 add_oil)
    present = {i['menu'] for i in inst}
    cand = [s for s in SNACK_POOL if s[0] not in present]
    if kmax is not None or pmax is not None:
        cur = totals(inst)
        cand = [s for s in cand
                if (kmax is None or cur['K'] + s[2] <= kmax)
                and (pmax is None or cur['P'] + s[3] <= pmax)]
    if not cand:
        return False
    menu = min(cand, key=lambda s: abs(s[1] - need_kcal))[0]   # 부족분에 가장 가까운 간식
    for it in expand([menu]):
        it['fixedK'] = True                   # 간식은 다른 레버가 또 건드리지 않게
        inst.append(it)
    return True


OIL_MENU_WORDS = ('나물', '무침', '볶음', '구이', '전')   # 참기름을 더해도 자연스러운 요리
OIL_ADD_MAX_G = 10.0  # 새로 추가하는 참기름은 요리 하나당 이 이하로(넉넉한 나물 수준, 88kcal 내외).
                       # 원래 5g로 잡았다가 열량 통과율이 너무 떨어져(-19) 10g으로 조정(2026-07-24, 사용자 확정).
VINEGAR_KW = ('식초',)  # 초무침·초절임류는 신맛+참기름의 고소함이 텁텁해져 안 어울림(2026-07-24, 사용자 지정)


def add_oil(inst, need_kcal):
    """열량 부족분을 기름으로 보충. '열량보충(기름)' 같은 별도 항목으로 덜렁 붙이지 않고
    (1) 이미 기름을 쓰는 요리의 기름을 늘리거나 (2) 나물·무침 같은 요리에 참기름을 더한다.
    5kcal 미만은 무시 — 예전엔 0.0g짜리 빈 항목이 25% 끼에 붙었다.
    다 못 채워도(반환 False) 호출부(lever_calorie)는 반환값을 안 보고 실제 총열량을 다시 재서
    모자라면 add_snack(과일)으로 이어받으므로 안전하다 — 억지로 한 접시에 몰아넣지 않아도 됨."""
    if need_kcal < 5:
        return True
    add_g = need_kcal / 8.84

    # (1) 이미 기름 쓰는 요리의 기름 증량 (원본 3배 한도 안에서 나눠 담음)
    room = []
    for i in inst:
        if i['group'] == '유지류' and i['E']:
            o = i.get('orig', i['amt']) or i['amt']
            r = max(0.0, o * OIL_CEIL - i['amt'])
            if r > 0:
                room.append((i, r))
    if room:
        tot = sum(r for _, r in room)
        give = min(add_g, tot)
        for i, r in room:
            i['amt'] += give * (r / tot)
        add_g -= give
        if add_g < 0.1:
            return True

    # (2) 남은 부족분은 나물·무침 등 '아직 기름이 없고 식초도 안 든' 요리에 참기름을 새로 추가.
    # 이미 기름이 있는 메뉴는 건너뛴다 — (1)단계에서 이미 상한(OIL_CEIL 3배)까지 늘려봤을 것이므로,
    # 여기서 같은 메뉴에 또 넣으면 상한을 우회해서 참기름이 두 줄로 중복되는 버그가 있었음
    # (2026-07-24 실측: 6g+9g=15g로 원본의 7.5배까지 늘어남).
    # 한 요리당 OIL_ADD_MAX_G까지만 넣고, 그래도 남으면 다른 오일-프리 요리로 넘어간다(몰아넣지 않음).
    ing_nut = NUT[0]
    d = ing_nut.get('참기름', {'E': 884, 'protein': 0, 'P': 0, 'K': 0, 'Na': 0, 'group': '유지류'})
    has_oil = {i['menu'] for i in inst if i['group'] == '유지류'}
    for menu in dict.fromkeys(i['menu'] for i in inst):
        if add_g < 0.1:
            break
        if menu in has_oil:
            continue
        if any(k in i['ing'] for i in inst if i['menu'] == menu for k in VINEGAR_KW):
            continue
        if any(w in menu for w in OIL_MENU_WORDS):
            give = min(add_g, OIL_ADD_MAX_G)
            inst.append({'menu': menu, 'ing': '참기름', 'amt': round(give, 1), 'orig': 0,
                         'E': d.get('E') or 884, 'protein': 0, 'P': d.get('P') or 0,
                         'K': d.get('K') or 0, 'Na': d.get('Na') or 0,
                         'group': '유지류', 'fixedK': True})
            has_oil.add(menu)
            add_g -= give
    return add_g < 0.1


def _pick_pool(pool, anchor, need):
    """앵커 후순위: 비앵커만으로 need(kcal)를 감당할 수 있으면 비앵커만 반환."""
    non = [i for i in pool if i['menu'] != anchor]
    cur_non = sum(i['amt'] / 100 * i['E'] for i in non)
    if cur_non > 0 and cur_non - need >= cur_non * RICE_FLOOR:
        return non, cur_non
    return pool, sum(i['amt'] / 100 * i['E'] for i in pool)


def lever_calorie(inst, lo, hi, anchor=None, allow_snack=True, kmax=None, pmax=None):
    """열량을 목표 범위로. 조절 대상 = 밥계열 + 유지류(기름).
    ★ 비율(f=hi/e)로 곱하면 안 됨: 밥만 그 비율로 줄여도 총량은 hi에 안 닿는다
      (855kcal, 밥330 -> f=.819 -> 밥270 -> 총795, 여전히 초과). lever_protein처럼
      '초과분을 밥에서 덜어내는' 식으로 계산해야 함.
    앵커(유저메뉴가 덮밥·면인 경우)는 후순위: 비앵커 밥/기름만으로 감당되면 안 건드림.
    allow_snack=False면 간식(과일 등 K/P를 갖는 메뉴) 대신 기름으로만 채움 — adjust()의 반복 루프
    중간 패스에서 간식을 넣으면 그 K/P를 다음 패스의 칼륨/인 레버가 보고 다른 메뉴를 깎는 되먹임이
    생기므로(2026-07-23 확인), 마지막 패스에서만 True로 호출."""
    e = totals(inst)['E']
    if e > hi:
        # 초과분 축소 순서(2026-07-22 확정): 밥(주식) 먼저 → 하한(40%)까지 다 써도 남으면 기름.
        # 예전엔 밥+기름을 한 풀로 묶어 같은 비율로 동시에 깎아 우선순위가 없었음.
        # ※ 단계 사이에 totals()로 실측 재계산 — amt_floor_of(개별 50% 하한)가 RICE_FLOOR(40%)보다
        #   엄격해 실제 감소량이 의도치보다 작을 수 있어, 산술 장부(빼기)가 아닌 실측으로 남은 초과분을 구함.
        rice_pool, rice_cur = _pick_pool([i for i in inst if i['menu'] in RICE and i['E']], anchor, e - hi)
        if rice_cur > 0:
            new = max(rice_cur - (e - hi), rice_cur * RICE_FLOOR)
            f = new / rice_cur
            for i in rice_pool:
                i['amt'] = max(i['amt'] * f, amt_floor_of(i))
        e2 = totals(inst)['E']
        if e2 > hi:
            oil_pool, oil_cur = _pick_pool([i for i in inst if i['group'] == '유지류' and i['E']], anchor, e2 - hi)
            if oil_cur > 0:
                new = max(oil_cur - (e2 - hi), 0)
                f = new / oil_cur
                for i in oil_pool:
                    i['amt'] = max(i['amt'] * f, amt_floor_of(i))
    elif e < lo and e > 0:
        rice = [i for i in inst if i['menu'] in RICE and i['E']]
        non = [i for i in rice if i['menu'] != anchor]
        pool = non if non else rice                       # 증량도 비앵커 밥 우선
        cur = sum(i['amt'] / 100 * i['E'] for i in pool)
        if cur > 0:
            new = min(cur + (lo - e), cur * 1.3)          # 부족분만큼 더함(밥 최대 1.3배)
            f = new / cur
            for i in pool:
                i['amt'] *= f
        e2 = totals(inst)['E']
        if e2 < lo:
            # 남은 부족분은 (1)기존 요리의 기름 증량 먼저 최대한 시도 (2)그래도 못 채운 나머지만
            # 간식(과일 등 K/P 있는 메뉴)으로 보충. 기름이 K=P=0이라 부작용이 없으므로 항상 우선
            # 시도하고, 간식은 기름으로도 못 메운 진짜 마지막 순간에만 최소한으로 투입(2026-07-23).
            add_oil(inst, lo - e2)
            e3 = totals(inst)['E']
            if e3 < lo and allow_snack:
                add_snack(inst, lo - e3, kmax=kmax, pmax=pmax)


def lever_calorie_capped(inst, lo, hi, anchor=None, allow_snack=True, kmax=None, pmax=None):
    """lever_calorie의 raw P 예산 cap 버전(2026-07-27, 두부·콩류 앵커 전용 조건부 경로).
    감소 분기(e>hi)와 add_oil/add_snack 경로는 lever_calorie와 완전히 동일(무변경) — cap은
    증량 분기(e<lo, 밥 증량)에만 적용. add_oil이 넣는 참기름/콩기름은 원본 데이터상 P=0이라
    cap이 자연히 무제한이고(밀도 0), add_snack은 자체적으로 kmax/pmax(=여기 pmax와 동일 값)로
    raw P 절대예산을 이미 체크하므로 별도 cap이 불필요 — 실험(protein_phosphorus_final_compare/
    cross_anchor)에서 이 가정을 실측으로 확인함."""
    e = totals(inst)['E']
    if e > hi:
        rice_pool, rice_cur = _pick_pool([i for i in inst if i['menu'] in RICE and i['E']], anchor, e - hi)
        if rice_cur > 0:
            new = max(rice_cur - (e - hi), rice_cur * RICE_FLOOR)
            f = new / rice_cur
            for i in rice_pool:
                i['amt'] = max(i['amt'] * f, amt_floor_of(i))
        e2 = totals(inst)['E']
        if e2 > hi:
            oil_pool, oil_cur = _pick_pool([i for i in inst if i['group'] == '유지류' and i['E']], anchor, e2 - hi)
            if oil_cur > 0:
                new = max(oil_cur - (e2 - hi), 0)
                f = new / oil_cur
                for i in oil_pool:
                    i['amt'] = max(i['amt'] * f, amt_floor_of(i))
    elif e < lo and e > 0:
        rice = [i for i in inst if i['menu'] in RICE and i['E']]
        non = [i for i in rice if i['menu'] != anchor]
        pool = non if non else rice
        cur = sum(i['amt'] / 100 * i['E'] for i in pool)
        if cur > 0:
            new = min(cur + (lo - e), cur * 1.3)
            ratio = new / cur
            if ratio > 1.0 + 1e-9:
                current_raw_P = totals(inst)['P']
                headroom = pmax - current_raw_P
                allowed_increase = max(0.0, headroom * PROTEIN_CALORIE_CAP_FRAC)
                P_pool = sum(i['amt'] / 100 * i['P'] for i in pool if i['P'] is not None)
                Amt_pool = sum(i['amt'] for i in pool)
                if Amt_pool <= 0:
                    final_ratio = ratio
                else:
                    requested_grams = Amt_pool * (ratio - 1.0)
                    density = P_pool / Amt_pool
                    allowed_grams = float('inf') if density <= 1e-9 else allowed_increase / density
                    applied_grams = max(0.0, min(requested_grams, allowed_grams))
                    final_ratio = 1.0 + applied_grams / Amt_pool
            else:
                final_ratio = ratio
            for i in pool:
                i['amt'] *= final_ratio
        e2 = totals(inst)['E']
        if e2 < lo:
            add_oil(inst, lo - e2)
            e3 = totals(inst)['E']
            if e3 < lo and allow_snack:
                add_snack(inst, lo - e3, kmax=kmax, pmax=pmax)


def _plant_protein_path_needed(inst, menus, anchor):
    """주찬(menus[2]) 또는 지정앵커가 두부·콩류 식품군인지 판정(2026-07-27).
    메뉴명 하드코딩 없이 재료 group='두류' 분류만 사용 — 앵커/주찬 메뉴의 정체성 재료
    (이름에 재료명이 들어간 것, 없으면 최대량 고형재료)가 '두류'군이면 True.
    True면 adjust()가 lever_phosphorus_rawP/lever_protein_capped/lever_calorie_capped
    (raw P 통일 + 증량 90% cap) 경로를 쓰고, False면 기존 레버 그대로 쓴다.
    채택 근거: protein_phosphorus_cross_anchor 교차앵커 검증에서 이 경로가 두부·콩류
    앵커에서만 뚜렷한 서비스지표 개선을 보였고 생선구이·육류에서는 효과가 미미해
    두부·콩류 전용으로 조건부 반영함(2026-07-27 확정)."""
    candidates = []
    if menus and len(menus) > 2 and menus[2] not in candidates:
        candidates.append(menus[2])
    if anchor and anchor not in candidates:
        candidates.append(anchor)
    for menu_name in candidates:
        items = [i for i in inst if i['menu'] == menu_name]
        if not items:
            continue
        identity_items = [i for i in items if is_identity_ingredient(i)]
        check_items = identity_items if identity_items else items
        main_item = max(check_items, key=lambda x: x['amt'])
        if main_item.get('group') == '두류':
            return True
    return False


def adjust(menus, b, anchor=None):
    """b = dict(Elo,Ehi,Plo,Phi,Kmax,Pmax,Namax,Na_total_target). 한 끼 단위.
    anchor = 유저 지정메뉴(재료 보존). 반환: before, after, inst, p_ok(인 충족 성공여부)."""
    inst = expand(menus)
    SWAP_LOG.clear()
    before = totals(inst)
    na_target = b.get('Na_total_target', NA_TOTAL_MEAL)   # 남은예산 이월(없으면 기본 655)
    lever_kimchi(inst)              # 김치 반찬 → 저염김치 (나트륨)
    lever_sodium(inst)              # 조미료 → 첨가염 소금1g(393)까지 적응적 축소
    lever_sodium_extra(inst, na_target)  # 고나트륨 메뉴 통째 축소 → 총나트륨 남은예산으로. 판정은 조미료만.
    p_ok = True
    # 두부·콩류 앵커(주찬 또는 지정메뉴) 전용 raw P 통일 경로(2026-07-27 조건부 반영).
    # 교차앵커 실험 결과 이 경로는 두부·콩류에서만 뚜렷이 개선하고 생선구이·육류에는 효과가
    # 미미해, 그 두 앵커는 완전히 기존 경로(변경 없음)를 그대로 탄다 — _plant_protein_path_needed
    # 주석 및 protein_phosphorus_cross_anchor_report.md 참고.
    use_rawP_path = _plant_protein_path_needed(inst, menus, anchor)
    # 레버 2회 반복(칼륨교체가 인을↑, 단백질스케일이 칼륨/인을↑ 되흔들어 1패스로 수렴 안 함).
    # 루프를 더 늘려보니: 끼단위 통과는 오르나 하루누적(남은예산 방식)은 오히려 내려감(72.9→70.7).
    # 진동도 무해 확인(루프 중 최선 ≈ 수렴값). 주지표=하루5영양이므로 2회가 최적.
    for pass_i in range(2):
        # 모든 레버가 앵커(유저 지정메뉴)를 후순위로 다룸: 비앵커로 목표 달성 가능하면 앵커는 안 건드림.
        # (금지가 아니라 후순위 -> 영양 통과 능력은 그대로, 앵커는 가능할 때만 보존)
        # 예외: lever_kimchi는 의도적으로 앵커 무시(저염김치 강제가 임상 우선).
        lever_potassium(inst, b['Kmax'], anchor=anchor)
        if use_rawP_path:
            p_ok = lever_phosphorus_rawP(inst, b['Pmax'], anchor=anchor, plo=b['Plo'])
        else:
            p_ok = lever_phosphorus(inst, b['Pmax'], anchor=anchor, plo=b['Plo'])
        if use_rawP_path:
            lever_protein_capped(inst, b['Plo'], b['Phi'], anchor=anchor, pmax=b['Pmax'])
        else:
            lever_protein(inst, b['Plo'], b['Phi'], anchor=anchor)
        # 조미료 재적용(단백질 레버가 간장 든 메뉴를 스케일해 되살린 것 정리)은 반드시 열량 레버 "앞".
        # 조미료엔 열량이 있어서(설탕·고추장·마요네즈), 열량을 lo에 딱 맞춘 뒤 조미료를 깎으면
        # 그대로 미달로 떨어짐. 열량이 항상 마지막이어야 함.
        lever_sodium(inst)
        # 단백질 레버가 고나트륨 메뉴(미역국·해산물볶음)를 되불릴 수 있어 총 나트륨 축소도 재적용.
        # 열량 레버 앞: 메뉴를 줄이면 열량이 떨어지므로 그 뒤 열량 레버가 밥/기름으로 보충.
        lever_sodium_extra(inst, na_target)
        # 간식(과일 등 K/P 있는 메뉴)은 마지막 패스에서만 허용 — 중간 패스에서 넣으면 다음 패스의
        # 칼륨/인 레버가 그 K/P를 보고 되먹임(다른 메뉴 깎기)을 일으킴(2026-07-23 확인).
        if use_rawP_path:
            lever_calorie_capped(inst, b['Elo'], b['Ehi'], anchor=anchor, allow_snack=(pass_i == 1),
                                  kmax=b['Kmax'], pmax=b['Pmax'])
        else:
            lever_calorie(inst, b['Elo'], b['Ehi'], anchor=anchor, allow_snack=(pass_i == 1),
                          kmax=b['Kmax'], pmax=b['Pmax'])
    return before, totals(inst), inst, p_ok


def standard_weight(height_cm, sex):
    """표준체중(kg) = 키(m)^2 × 계수 (남 22 / 여 21).
    투석 영양 목표는 '실제 체중'이 아니라 이 표준체중을 기준으로 잡아야 한다(임상 지침).
    sex: 여성 계열('여','여성','f','female','w')이면 21, 그 외는 남성 22."""
    h = height_cm / 100.0
    female = str(sex).strip().lower() in ('여', '여성', 'f', 'female', 'w', 'women')
    return h * h * (21 if female else 22)


def day_targets(weight):
    # 하루 기준. weight는 '표준체중'(standard_weight)을 넣는다.
    # 칼륨·인은 '미만'(<) 기준 — day_ok/passes에서 < 로 판정. 나트륨은 조미료 끼니고정(소금1g x 3).
    return {'Elo': weight * 30, 'Ehi': weight * 35,
            'Plo': weight * 1.1, 'Phi': weight * 1.2,
            'Kmax': 3000, 'Pmax': 1000, 'Namax': SALT_MG * 3}


def meal_bounds(weight, consumed=None, meals_left=3, total_meals=3, tol=0.2):
    """남은 예산 + 균형 밴드 방식. 한 끼 목표 반환.
    consumed   : 오늘 이미 먹은 누적 {'E','protein','K','P','Na'} (첫 끼면 None)
    meals_left : 이번 끼 포함 남은 끼니 수 (아침=3, 점심=2, 저녁=1)
    tol        : 1일÷3 대비 허용 편차(몰아먹기 방지 밴드). 0.2 = ±20%
    나트륨: 첨가염(Namax)은 끼니 고정 393mg — 조미료는 맛 보존 하한이 있어 예산화 의미 없음.
    단, 총나트륨(자연재료 포함)은 Na_total_target로 남은예산÷남은끼니 이월(K·P와 동일 cap() 방식).
    앞 끼에 미역국 등으로 자연나트륨을 많이 먹었으면 다음 끼 목표가 빡빡해져 lever_sodium_extra가 더 세게 줄인다.
    """
    d = day_targets(weight)
    c = consumed or {'E': 0, 'protein': 0, 'K': 0, 'P': 0, 'Na': 0}
    n = max(1, meals_left)

    def rng(dv, used):   # 범위형(열량·단백질) 하한/상한: 밴드로 클램프
        fair = dv / total_meals
        return min(max((dv - used) / n, fair * (1 - tol)), fair * (1 + tol))

    def cap(dv, used):   # 상한형(칼륨·인·총나트륨): 남은예산은 아래로 자유, 위는 가드레일
        fair = dv / total_meals
        return min((dv - used) / n, fair * (1 + tol))

    return {'Elo': rng(d['Elo'], c['E']), 'Ehi': rng(d['Ehi'], c['E']),
            'Plo': rng(d['Plo'], c['protein']), 'Phi': rng(d['Phi'], c['protein']),
            'Kmax': cap(d['Kmax'], c['K']),
            # 인(P)은 raw P가 하드 기준(passes()와 동일) — 남은 예산도 반드시 raw P로 이월한다.
            # Peff(흡수보정 인)는 최적화용 소프트 신호일 뿐 예산·상한 판정에는 쓰지 않는다.
            'Pmax': cap(d['Pmax'], c.get('P', 0)),
            'Namax': SALT_MG,   # 나트륨 끼니 고정 = 조미료 소금 1g (자연염 제외)
            'Na_total_target': cap(NA_TOTAL_MEAL * total_meals, c.get('Na', 0))}   # 총나트륨 남은예산


def main():
    global NUT
    NUT = load_all()
    rows = list(csv.reader(open('data/FOOK_diet_1_12_kor.csv', encoding='cp949')))
    W = 60
    dt = day_targets(W)

    EPS = 1e-6   # 부동소수점 누적오차 허용오차 — 예: 1799.9999999999998가 정확히 1800인데 미달로 오판되는 것 방지
    def day_ok(t):   # 하루 총합 판정 (나트륨은 조미료 끼니고정 393x3=1179 상한)
        return (dt['Elo'] - EPS <= t['E'] <= dt['Ehi'] + EPS, dt['Plo'] - EPS <= t['protein'] <= dt['Phi'] + EPS,
                t['K'] < dt['Kmax'] + EPS, t['P'] < dt['Pmax'] + EPS, t['Na_season'] <= dt['Namax'] + EPS)

    labels = ['열량', '단백질', '칼륨', '인', '나트륨']
    n_days = 0
    cnt = [0] * 5
    all_ok = 0
    meal_na_ok = 0
    meal_na_warn = 0   # 첨가염은 정상인데 총나트륨(자연재료 포함)이 786(소금2g) 넘는 끼니 — 별도 지표
    n_meals = 0
    for r in rows[1:]:
        slots = [r[1:6], r[6:11], r[11:16]]   # 아침/점심/저녁
        meals = [[c.strip() for c in sl if c.strip() and c.strip() != 'empty'] for sl in slots]
        if not all(meals):
            continue
        consumed = {'E': 0, 'protein': 0, 'K': 0, 'P': 0, 'Peff': 0, 'Na': 0, 'Na_season': 0}
        for mi, m in enumerate(meals):
            b = meal_bounds(W, consumed, meals_left=3 - mi)   # 남은예산: 누적 반영
            _, at, _, _ = adjust(m, b)
            for k in consumed:
                consumed[k] = round(consumed[k] + at[k], 4)   # 부동소수점 누적오차 방지(방법2)
            n_meals += 1
            season_ok = at['Na_season'] <= b['Namax']
            if season_ok:
                meal_na_ok += 1
                if at['Na'] > NA_TOTAL_WARN:
                    meal_na_warn += 1   # 첨가염 제어는 성공, 자연나트륨(미역국·해산물 등)이 원인
        n_days += 1
        k = day_ok(consumed)
        for i in range(5):
            cnt[i] += k[i]
        all_ok += all(k)
    print(f'하루 기준 (체중 {W}kg): 열량 {dt["Elo"]:.0f}~{dt["Ehi"]:.0f}, 단백 {dt["Plo"]:.0f}~{dt["Phi"]:.0f}, '
          f'K<{dt["Kmax"]:.0f}, P<{dt["Pmax"]:.0f}  |  나트륨 끼니 786mg 고정  (총 {n_days}일)')
    print(f'{"영양소":<6} {"하루충족":>10}')
    for i, l in enumerate(labels):
        print(f'  {l:<6} {cnt[i]:>6}/{n_days}')
    print(f'  하루 5개 전부 충족: {all_ok}/{n_days}')
    print(f'  끼니 첨가염 제어율: {meal_na_ok}/{n_meals} ({meal_na_ok/n_meals*100:.1f}%)  — 레버 성능 지표')
    print(f'  끼니 자연나트륨 경고율: {meal_na_warn}/{n_meals} ({meal_na_warn/n_meals*100:.1f}%)  — 첨가염은 정상, '
          f'재료 자체(미역국·해산물 등)가 원인. 강제축소 대신 경고 안내 대상')


if __name__ == '__main__':
    main()
